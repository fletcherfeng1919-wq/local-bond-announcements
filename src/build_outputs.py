"""Build the deliverables from data/state_announcements.csv and
data/state_plans.csv:
  output/local_raw_data.xlsx  (原始全量明细 + 公式校验版 + 月度季度计划明细)
  output/local_summary.xlsx   (按期限/省份/品种的分组统计 + 前置周期 + 异常值清单)
  charts/boxplot_workday_gap_by_term.png
"""
import datetime

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
matplotlib.rcParams["font.sans-serif"] = ["PingFang SC", "Heiti SC", "Arial Unicode MS", "DejaVu Sans"]
matplotlib.rcParams["axes.unicode_minus"] = False
import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from . import config
from .pipeline import load_state

TERM_ORDER = config.VALID_TERMS
CATEGORY_ORDER = [config.CATEGORY_NEW_GENERAL, config.CATEGORY_NEW_SPECIAL, config.CATEGORY_REFINANCING]

ANN_COLUMN_LABELS = {
    "title": "公告网页标题",
    "pub_date": "公告发布日期",
    "url": "公告链接",
    "source_name": "数据来源栏目",
    "doc_type": "文件类型",
    "province": "所属省份/计划单列市",
    "province_code": "省份代码",
    "category_code": "债券类型代码",
    "category_label": "债券类型",
    "category_subtype": "再融资子类(一般/专项)",
    "term": "发行期限",
    "batch_no": "批次(标题)",
    "issue_no": "期数",
    "issue_no_range": "期数范围(合并行)",
    "total_amount_yi": "发行规模(亿元)",
    "bid_date": "招投标日期(基准发行日)",
    "base_date_type": "基准日口径",
    "payment_date": "起息日",
    "listing_date": "上市流通日",
    "natural_day_gap": "自然日间隔",
    "workday_gap": "工作日间隔",
    "doc_no": "发文号",
    "bond_name": "债券名称(PDF表格原文)",
    "extraction_method": "提取方式(text/ocr/failed)",
    "warnings": "提取备注/警告",
}

PLAN_COLUMN_LABELS = {
    "title": "公告网页标题",
    "pub_date": "计划公示日期",
    "url": "公告链接",
    "source_name": "数据来源栏目",
    "doc_type": "文件类型",
    "province": "所属省份/计划单列市",
    "province_code": "省份代码",
    "covered_year": "覆盖年份",
    "covered_month_start": "覆盖起始月",
    "covered_month_end": "覆盖结束月",
    "covered_period_start": "覆盖期起始日",
    "covered_period_end": "覆盖期结束日",
    "plan_general_amount_yi": "计划新增一般债券规模(亿元)",
    "plan_special_amount_yi": "计划新增专项债券规模(亿元)",
    "plan_refinancing_amount_yi": "计划再融资债券规模(亿元)",
    "extraction_method": "提取方式(text/ocr/failed)",
    "warnings": "提取备注/警告",
}

RESULT_COLUMN_LABELS = {
    "title": "公告网页标题",
    "pub_date": "结果公开日期",
    "url": "公告链接",
    "source_name": "数据来源栏目",
    "doc_type": "文件类型",
    "province": "所属省份/计划单列市",
    "province_code": "省份代码",
    "bond_name": "债券名称",
    "bond_code": "债券编码",
    "bond_short_name": "债券简称(市场代码)",
    "bond_market_type": "债券类型(原文)",
    "category_code": "债券类型代码",
    "category_label": "债券类型",
    "term": "发行期限",
    "total_amount_yi": "发行规模(亿元)",
    "new_amount_yi": "新增债券(亿元)",
    "swap_amount_yi": "置换债券(亿元)",
    "refi_amount_yi": "再融资债券(亿元)",
    "batch_label": "发行批次",
    "coupon_rate_pct": "票面利率(%)",
    "issue_date": "发行日期",
    "value_date": "起息日",
    "payment_freq": "付息方式",
    "redemption_structure": "赎回模式",
    "extraction_method": "提取方式(text/ocr/unsupported_legacy_format)",
    "warnings": "提取备注/警告",
}

RESULT_EXPORT_COLS = [
    "bond_code", "bond_short_name", "bond_name", "province", "category_label",
    "term", "total_amount_yi", "new_amount_yi", "swap_amount_yi", "refi_amount_yi",
    "coupon_rate_pct", "issue_date", "value_date", "payment_freq", "redemption_structure",
    "batch_label", "bond_market_type", "pub_date", "doc_type", "source_name",
    "extraction_method", "url", "warnings",
]

ANN_EXPORT_COLS = [
    "title", "pub_date", "province", "category_label", "category_subtype", "term",
    "batch_no", "issue_no", "issue_no_range", "total_amount_yi", "bid_date",
    "base_date_type", "payment_date", "listing_date", "natural_day_gap", "workday_gap",
    "doc_no", "bond_name", "doc_type", "source_name", "extraction_method", "url", "warnings",
]

PLAN_EXPORT_COLS = [
    "title", "pub_date", "province", "covered_year", "covered_month_start",
    "covered_month_end", "covered_period_start", "covered_period_end",
    "plan_general_amount_yi", "plan_special_amount_yi", "plan_refinancing_amount_yi",
    "doc_type", "source_name", "extraction_method", "url", "warnings",
]


def _prep_announcements() -> pd.DataFrame:
    df = load_state("announcements")
    df["term"] = pd.Categorical(df["term"], categories=TERM_ORDER, ordered=True)
    df["category_code"] = pd.Categorical(df["category_code"], categories=CATEGORY_ORDER, ordered=True)
    return df


def _prep_plans() -> pd.DataFrame:
    return load_state("plans")


def _prep_results() -> pd.DataFrame:
    df = load_state("results")
    df["term"] = pd.Categorical(df["term"], categories=TERM_ORDER, ordered=True)
    return df


def _autosize(ws, cols, labels):
    for i, col in enumerate(cols, start=1):
        header = labels.get(col, col)
        ws.column_dimensions[get_column_letter(i)].width = min(max(12, len(header) * 2 + 4), 45)


def _bold_header(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True)


# ---------------------------------------------------------------------------
# local_raw_data.xlsx
# ---------------------------------------------------------------------------

def build_raw_data_xlsx(ann_df: pd.DataFrame, plan_df: pd.DataFrame):
    out_path = config.OUTPUT_DIR / "local_raw_data.xlsx"

    ann_export = ann_df[ANN_EXPORT_COLS].copy()
    ann_export = ann_export.astype(object).where(ann_export.notna(), None)
    ann_export.columns = [ANN_COLUMN_LABELS[c] for c in ANN_EXPORT_COLS]

    plan_export = plan_df[PLAN_EXPORT_COLS].copy() if len(plan_df) else pd.DataFrame(columns=PLAN_EXPORT_COLS)
    plan_export = plan_export.astype(object).where(plan_export.notna(), None)
    plan_export.columns = [PLAN_COLUMN_LABELS[c] for c in PLAN_EXPORT_COLS]

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        ann_export.to_excel(writer, sheet_name="原始全量明细(发行前公告)", index=False)
        plan_export.to_excel(writer, sheet_name="月度季度发行计划明细", index=False)

    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    ws1 = wb["原始全量明细(发行前公告)"]
    _bold_header(ws1)
    _autosize(ws1, ANN_EXPORT_COLS, ANN_COLUMN_LABELS)

    ws_plan = wb["月度季度发行计划明细"]
    _bold_header(ws_plan)
    _autosize(ws_plan, PLAN_EXPORT_COLS, PLAN_COLUMN_LABELS)

    # --- formula-verification sheet: recompute natural/workday gap live in
    # Excel via NETWORKDAYS, so the two Python-computed gap columns on sheet 1
    # can be cross-checked without trusting the Python computation blindly.
    ws2 = wb.create_sheet("公式校验版")
    header = [ANN_COLUMN_LABELS[c] for c in ANN_EXPORT_COLS]
    ws2.append(header)
    _bold_header(ws2)

    pub_col = ANN_EXPORT_COLS.index("pub_date") + 1
    bid_col = ANN_EXPORT_COLS.index("bid_date") + 1
    natday_col = ANN_EXPORT_COLS.index("natural_day_gap") + 1
    workday_col = ANN_EXPORT_COLS.index("workday_gap") + 1

    if len(ann_df) and ann_df["pub_date"].notna().any():
        lo = min(ann_df["pub_date"].dropna().min(), ann_df["bid_date"].dropna().min()) - datetime.timedelta(days=10)
        hi = max(ann_df["pub_date"].dropna().max(), ann_df["bid_date"].dropna().max()) + datetime.timedelta(days=30)
    else:
        lo, hi = datetime.date(2015, 1, 1), datetime.date(2026, 12, 31)

    holiday_dates, makeup_workday_dates = [], []
    from chinese_calendar import is_workday
    d = lo
    while d <= hi:
        try:
            wd = is_workday(d)
        except NotImplementedError:
            d += datetime.timedelta(days=1)
            continue
        if d.weekday() < 5 and not wd:
            holiday_dates.append(d)
        if d.weekday() >= 5 and wd:
            makeup_workday_dates.append(d)
        d += datetime.timedelta(days=1)

    helper_ws = wb.create_sheet("法定节假日辅助表")
    helper_ws.append(["法定节假日(工作日调休为休息)", "调休补班日(周末调整为工作日)"])
    _bold_header(helper_ws)
    for i in range(max(len(holiday_dates), len(makeup_workday_dates))):
        helper_ws.append([
            holiday_dates[i] if i < len(holiday_dates) else None,
            makeup_workday_dates[i] if i < len(makeup_workday_dates) else None,
        ])
    for c in ("A", "B"):
        helper_ws.column_dimensions[c].width = 28
    n_holidays, n_makeup = len(holiday_dates), len(makeup_workday_dates)

    for i, row in enumerate(ann_export.itertuples(index=False), start=2):
        ws2.append(list(row))
        pub_cell = f"{get_column_letter(pub_col)}{i}"
        bid_cell = f"{get_column_letter(bid_col)}{i}"
        ws2[f"{get_column_letter(natday_col)}{i}"] = (
            f"=IF(OR({bid_cell}=\"\",{pub_cell}=\"\"),\"\",{bid_cell}-{pub_cell})"
        )
        holiday_rng = f"法定节假日辅助表!$A$2:$A${1 + n_holidays}" if n_holidays else None
        makeup_rng = f"法定节假日辅助表!$B$2:$B${1 + n_makeup}" if n_makeup else None
        if holiday_rng and makeup_rng:
            formula = (
                f"=IF(OR({bid_cell}=\"\",{pub_cell}=\"\"),\"\","
                f"NETWORKDAYS({pub_cell}+1,{bid_cell},{holiday_rng})"
                f"+COUNTIFS({makeup_rng},\">=\"&({pub_cell}+1),{makeup_rng},\"<=\"&{bid_cell}))"
            )
        else:
            formula = f"=IF(OR({bid_cell}=\"\",{pub_cell}=\"\"),\"\",NETWORKDAYS({pub_cell}+1,{bid_cell}))"
        ws2[f"{get_column_letter(workday_col)}{i}"] = formula

    _autosize(ws2, ANN_EXPORT_COLS, ANN_COLUMN_LABELS)

    for ws in (ws1, ws2):
        for col in ("pub_date", "bid_date", "payment_date", "listing_date"):
            idx = ANN_EXPORT_COLS.index(col) + 1
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=idx).number_format = "yyyy-mm-dd"
    for col in ("pub_date", "covered_period_start", "covered_period_end"):
        idx = PLAN_EXPORT_COLS.index(col) + 1
        for r in range(2, ws_plan.max_row + 1):
            ws_plan.cell(row=r, column=idx).number_format = "yyyy-mm-dd"

    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# local_bond_code_dictionary.xlsx
# ---------------------------------------------------------------------------

def build_bond_dictionary_xlsx(results_df: pd.DataFrame):
    """The 发行结果-derived bond_code/bond_short_name lookup table -- one row
    per bond, usable to join this project's other tables (or Wind/Bloomberg
    pulls) onto an actual market code instead of a free-text bond name."""
    out_path = config.OUTPUT_DIR / "local_bond_code_dictionary.xlsx"

    coded = results_df[results_df["bond_code"].notna()].copy()
    legacy = results_df[results_df["extraction_method"] == "unsupported_legacy_format"].copy()

    coded_export = coded[RESULT_EXPORT_COLS].copy()
    coded_export = coded_export.astype(object).where(coded_export.notna(), None)
    coded_export.columns = [RESULT_COLUMN_LABELS[c] for c in RESULT_EXPORT_COLS]

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        coded_export.to_excel(writer, sheet_name="债券代码字典", index=False)
        if len(legacy):
            legacy_export = legacy[["title", "pub_date", "province", "url"]].copy()
            legacy_export.columns = ["公告网页标题", "结果公开日期", "所属省份", "公告链接"]
            legacy_export.to_excel(writer, sheet_name="2020年以前(暂无代码)", index=False)

    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    for ws in wb.worksheets:
        _bold_header(ws)
        for i, col in enumerate(ws.iter_cols(1, ws.max_column), start=1):
            header_len = len(str(col[0].value or ""))
            ws.column_dimensions[get_column_letter(i)].width = max(14, header_len + 4)
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# local_summary.xlsx
# ---------------------------------------------------------------------------

def _lead_stats(g: pd.DataFrame) -> pd.Series:
    s = g["natural_day_gap"].dropna()
    ws = g["workday_gap"].dropna()
    return pd.Series({
        "样本数量": len(g),
        "平均提前自然日": round(s.mean(), 2) if len(s) else None,
        "中位数提前自然日": s.median() if len(s) else None,
        "最大提前自然日": s.max() if len(s) else None,
        "最小提前自然日": s.min() if len(s) else None,
        "平均提前工作日": round(ws.mean(), 2) if len(ws) else None,
        "中位数提前工作日": ws.median() if len(ws) else None,
        "最大提前工作日": ws.max() if len(ws) else None,
        "最小提前工作日": ws.min() if len(ws) else None,
    })


def _plan_lead_join(ann_df: pd.DataFrame, plan_df: pd.DataFrame) -> pd.DataFrame:
    """For each monthly/quarterly plan row, find actual bid dates observed in
    the same province whose bid_date falls inside the plan's covered period,
    and summarize the gap from the plan's own publish date to those actual
    bid dates -- this is the '计划公示日距离实际招投标日期' lead time the
    plan PDF itself doesn't state directly (it only gives an aggregate
    monthly amount, not per-bond dates)."""
    from .workdays import workday_diff

    rows = []
    ann_valid = ann_df.dropna(subset=["bid_date", "province"])
    for _, plan in plan_df.iterrows():
        if pd.isna(plan.get("covered_period_start")) or pd.isna(plan.get("pub_date")):
            continue
        match = ann_valid[
            (ann_valid["province"] == plan["province"])
            & (ann_valid["bid_date"] >= plan["covered_period_start"])
            & (ann_valid["bid_date"] <= plan["covered_period_end"])
        ]
        if match.empty:
            continue
        gaps_natural = [(bd - plan["pub_date"]).days for bd in match["bid_date"]]
        gaps_workday = [workday_diff(plan["pub_date"], bd) for bd in match["bid_date"]]
        gaps_workday = [g for g in gaps_workday if g is not None]
        rows.append({
            "省份": plan["province"], "计划公示日期": plan["pub_date"],
            "覆盖期间": f"{plan['covered_period_start']}~{plan['covered_period_end']}",
            "匹配到的实际发行笔数": len(match),
            "最短提前自然日": min(gaps_natural), "最长提前自然日": max(gaps_natural),
            "平均提前自然日": round(sum(gaps_natural) / len(gaps_natural), 1),
            "最短提前工作日": min(gaps_workday) if gaps_workday else None,
            "最长提前工作日": max(gaps_workday) if gaps_workday else None,
            "平均提前工作日": round(sum(gaps_workday) / len(gaps_workday), 1) if gaps_workday else None,
        })
    return pd.DataFrame(rows)


WEEKDAY_CN = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]


def _add_derived_time_fields(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["region"] = df["province"].map(config.PROVINCE_TO_REGION)
    bid = pd.to_datetime(df["bid_date"])
    df["bid_weekday"] = bid.dt.weekday.map(lambda d: WEEKDAY_CN[int(d)] if pd.notna(d) else None)
    df["bid_year"] = bid.dt.year
    df["bid_quarter"] = bid.dt.year.astype("Int64").astype(str) + "Q" + bid.dt.quarter.astype("Int64").astype(str)
    return df


def build_summary_stat_xlsx(ann_df: pd.DataFrame, plan_df: pd.DataFrame):
    out_path = config.OUTPUT_DIR / "local_summary.xlsx"
    valid = ann_df.dropna(subset=["category_code"]).copy()
    valid = _add_derived_time_fields(valid)

    by_term_category = (
        valid.groupby(["category_label", "term"], observed=True)
        .apply(_lead_stats, include_groups=False).reset_index()
    )
    by_category = (
        valid.groupby("category_label", observed=True)
        .apply(_lead_stats, include_groups=False).reset_index()
    )
    by_province = (
        valid.groupby("province", observed=True)
        .apply(_lead_stats, include_groups=False).reset_index()
        .sort_values("样本数量", ascending=False)
    )
    by_province_category = (
        valid.groupby(["province", "category_label"], observed=True)
        .apply(_lead_stats, include_groups=False).reset_index()
    )
    by_region = (
        valid.dropna(subset=["region"]).groupby("region", observed=True)
        .apply(_lead_stats, include_groups=False).reset_index()
        .rename(columns={"region": "区域分组"})
    )
    by_region["区域分组"] = pd.Categorical(by_region["区域分组"], categories=config.REGION_ORDER, ordered=True)
    by_region = by_region.sort_values("区域分组")
    by_region_category = (
        valid.dropna(subset=["region"]).groupby(["region", "category_label"], observed=True)
        .apply(_lead_stats, include_groups=False).reset_index()
        .rename(columns={"region": "区域分组"})
    )
    by_weekday = (
        valid.dropna(subset=["bid_weekday"]).groupby("bid_weekday", observed=True)
        .apply(_lead_stats, include_groups=False).reset_index()
        .rename(columns={"bid_weekday": "招标日星期"})
    )
    by_weekday["招标日星期"] = pd.Categorical(by_weekday["招标日星期"], categories=WEEKDAY_CN, ordered=True)
    by_weekday = by_weekday.sort_values("招标日星期")
    by_quarter = (
        valid.dropna(subset=["bid_quarter"]).groupby("bid_quarter", observed=True)
        .apply(_lead_stats, include_groups=False).reset_index()
        .rename(columns={"bid_quarter": "招标日所属季度"})
        .sort_values("招标日所属季度")
    )

    anomalies = ann_df[ann_df["workday_gap"] < config.STATUTORY_MIN_WORKDAYS].copy()
    anomalies = anomalies.sort_values("workday_gap")

    plan_lead = _plan_lead_join(ann_df, plan_df)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        by_term_category.to_excel(writer, sheet_name="分组1_按期限x品种", index=False)
        by_category.to_excel(writer, sheet_name="分组2_按品种", index=False)
        by_province.to_excel(writer, sheet_name="分组3_按省份", index=False)
        by_province_category.to_excel(writer, sheet_name="分组4_按省份x品种", index=False)
        by_region.to_excel(writer, sheet_name="分组5_按区域(自定义)", index=False)
        by_region_category.to_excel(writer, sheet_name="分组6_按区域x品种", index=False)
        by_weekday.to_excel(writer, sheet_name="分组7_按招标日星期", index=False)
        by_quarter.to_excel(writer, sheet_name="分组8_按招标日季度(时间趋势)", index=False)
        plan_lead.to_excel(writer, sheet_name="月度季度计划前置周期", index=False)
        anomalies[ANN_EXPORT_COLS].rename(columns=ANN_COLUMN_LABELS).to_excel(
            writer, sheet_name=f"异常值_工作日间隔小于{config.STATUTORY_MIN_WORKDAYS}天", index=False
        )

    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    for ws in wb.worksheets:
        _bold_header(ws)
        for i, col in enumerate(ws.iter_cols(1, ws.max_column), start=1):
            header_len = len(str(col[0].value or ""))
            ws.column_dimensions[get_column_letter(i)].width = max(14, header_len + 4)
    wb.save(out_path)

    return out_path, {
        "by_term_category": by_term_category, "by_category": by_category,
        "by_province": by_province, "by_province_category": by_province_category,
        "by_region": by_region, "by_region_category": by_region_category,
        "by_weekday": by_weekday, "by_quarter": by_quarter,
        "plan_lead": plan_lead, "anomalies": anomalies,
    }


# ---------------------------------------------------------------------------
# Boxplot chart
# ---------------------------------------------------------------------------

BLUE = "#2a78d6"
INK = "#0b0b0b"
SECONDARY_INK = "#52514e"
MUTED = "#898781"
GRID = "#e1e0d9"
SURFACE = "#fcfcfb"


def build_boxplot(ann_df: pd.DataFrame):
    valid = ann_df.dropna(subset=["workday_gap", "term"]).copy()
    present_terms = [t for t in TERM_ORDER if (valid["term"] == t).sum() > 0]
    if not present_terms:
        return None

    data = [valid.loc[valid["term"] == t, "workday_gap"].values for t in present_terms]
    counts = [len(d) for d in data]

    fig, ax = plt.subplots(figsize=(11, 6), dpi=150)
    fig.patch.set_facecolor(SURFACE)
    ax.set_facecolor(SURFACE)

    bp = ax.boxplot(
        data, tick_labels=present_terms, patch_artist=True, widths=0.5,
        medianprops=dict(color=INK, linewidth=2),
        boxprops=dict(facecolor=BLUE, edgecolor=BLUE, alpha=0.35, linewidth=1.4),
        whiskerprops=dict(color=SECONDARY_INK, linewidth=1.2),
        capprops=dict(color=SECONDARY_INK, linewidth=1.2),
        flierprops=dict(marker="o", markerfacecolor=BLUE, markeredgecolor="none", markersize=4, alpha=0.5),
    )
    for patch in bp["boxes"]:
        patch.set_edgecolor(BLUE)

    ax.axhline(config.STATUTORY_MIN_WORKDAYS, color="#c0392b", linewidth=1, linestyle="--")
    ax.annotate(f"法定底线 {config.STATUTORY_MIN_WORKDAYS} 个工作日", (0.55, config.STATUTORY_MIN_WORKDAYS),
                xycoords=("axes fraction", "data"), fontsize=8, color="#c0392b", va="bottom")

    ax.set_title("各期限地方债：公告发布日至招投标日提前工作日分布", fontsize=13, color=INK, pad=14)
    ax.set_xlabel("发行期限", fontsize=10, color=SECONDARY_INK)
    ax.set_ylabel("提前工作日数", fontsize=10, color=SECONDARY_INK)
    ax.grid(axis="y", color=GRID, linewidth=0.8, zorder=0)
    ax.set_axisbelow(True)
    for spine in ("top", "right"):
        ax.spines[spine].set_visible(False)
    for spine in ("left", "bottom"):
        ax.spines[spine].set_color(MUTED)
    ax.tick_params(colors=SECONDARY_INK)

    for i, n in enumerate(counts, start=1):
        ax.annotate(f"n={n}", (i, ax.get_ylim()[1]), ha="center", va="bottom",
                    fontsize=8, color=MUTED, annotation_clip=False)

    fig.tight_layout()
    out_path = config.CHARTS_DIR / "boxplot_workday_gap_by_term.png"
    fig.savefig(out_path, facecolor=SURFACE)
    plt.close(fig)
    return out_path
