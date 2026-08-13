"""Three-way cross-validation: celma.org.cn (data/state_results.csv, the
primary source) vs SSE listing notices (src/sse_listing.py) vs provincial
finance-bureau announcements (src/provincial_crawl.py + provincial_verify.py).

Each source has a different reliability profile, so this module treats
them asymmetrically rather than pretending they're interchangeable:
  - **celma** is the base/primary -- everything is diffed AGAINST it, not
    merged with it as an equal.
  - **SSE** only confirms a bond EXISTS and is trading (securityAbbr,
    sseDate) -- no amount/rate. Its only useful signal here is "celma is
    missing a bond SSE already knows about" (a real, if partial, gap).
  - **Provincial** sources carry the fullest data (amount/term/rate) but
    are the least reliable end-to-end (OCR noise on scanned sources) --
    provincial_verify.diff_against_state()'s `low_confidence` bucket
    already screens out rows whose fields might be cross-contaminated;
    this module never treats anything in that bucket as evidence of a
    real discrepancy, and apply_corrections() never writes from it.

## Why corrections are applied more conservatively here than in wind_reconcile.py

wind_reconcile.py auto-applies amount/rate/date patches from a Wind
terminal export because that source is clean, authoritative, manually
verified data. Provincial sources are structurally noisier (most of this
project's own extraction bugs were discovered BECAUSE of exactly this kind
of scanned-PDF/OCR path -- see extract_result.py's history). So
apply_corrections() here only patches a row when ALL of:
  1. the provincial row is not in `low_confidence`
  2. it matched a UNIQUE candidate in state_results.csv (bond_short_name
     exact match, or a (province, issue_date, term) key that isn't
     ambiguous between multiple same-day same-term bonds)
  3. the mismatch is a genuine value difference, not just a field the
     provincial row failed to extract (None never overwrites a real value)
It never ADDS new rows the way wind_reconcile.py does for Wind-only bonds
-- provincial templates mostly lack bond_short_name/bond_code (see
provincial_verify.py's module docstring), so a "missing from celma" row
here can't be safely inserted without risking a duplicate once celma
eventually picks up the same bond under its own code. Those cases are
reported, not auto-applied.
"""
import datetime

import pandas as pd

from . import config, provincial_crawl, provincial_verify, sse_listing

AMOUNT_TOLERANCE_YI = provincial_verify.AMOUNT_TOLERANCE_YI
RATE_TOLERANCE_PCT = provincial_verify.RATE_TOLERANCE_PCT


def _months_back_start(months_back: int) -> str:
    d = datetime.date.today().replace(day=1)
    for _ in range(months_back - 1):
        d = (d - datetime.timedelta(days=1)).replace(day=1)
    return d.isoformat()


def check_sse_coverage(state_df: pd.DataFrame, months_back: int = 2) -> dict:
    """For every SSE 上市公告 in the trailing `months_back` window, check
    whether celma (state_results.csv) already has a matching bond_short_name.
    Returns {"checked": N, "missing_from_celma": [...], "missing_by_province": {...}}
    -- SSE has no amount/rate to cross-check, so this can only ever report
    existence gaps, never value mismatches.

    Confirmed (2026-08-12) SSE's `securityAbbr` and celma's `bond_short_name`
    are not always byte-identical for the same bond -- SSE drops the "债"
    character celma's short name carries (celma: "26江苏债31", SSE:
    "26江苏31"). A raw-string version of this comparison reported 107/156
    notices "missing from celma" in a 2-month window; manually checking one
    province's worth (江苏) showed all 6 of its "missing" bonds were sitting
    right there in state_results.csv under the celma-convention name.
    Matching on a "债"-stripped form of both sides collapsed that 107 down
    to 4 genuinely missing bonds (both in provinces -- 内蒙古/黑龙江 -- outside
    this project's 10-province coverage, both dated within the trailing
    week, consistent with celma's ordinary ~1-week lag). **Do not revert to
    a raw string comparison here** -- 107 looked like a real finding but was
    this project's own matching bug.

    The remaining (small, real) gap still concentrates heavily by PROVINCE,
    not smoothly by recency the way a pure "celma lags ~1 week" story would
    predict -- e.g. 青岛市 (Qingdao) showed 100% missing on a notice 3 weeks
    old, while other provinces from the same week showed 0% missing. This
    matches celma's own already-documented irregular publishing rhythm for
    certain issuers (青岛's confirmed-results announcements arrive in
    infrequent bursts covering several months at once, not steadily) --
    the per-province breakdown here is what makes that visible instead of
    a single flat count that looks like a uniform lag.

    **Second naming bug found 2026-08-13 (same family as the "债" one
    above, do not let this recur a third time)**: SSE abbreviates two
    provinces' ROOT names, not just the usual 省/自治区 suffix -- 黑龙江省
    becomes "龙江" (drops the leading char) and 内蒙古自治区 becomes "内蒙"
    (drops the trailing char of the root), while celma keeps the full root
    ("26黑龙江债22", "26内蒙古债18"). These are the only two provinces in
    config.PROVINCES whose root name is 3+ characters, so this is a closed,
    fully-enumerable set, not a pattern needing fuzzy matching. Caught by
    manually cross-referencing what looked like 4 "genuine" gaps (内蒙古×2,
    黑龙江×2) against state_results.csv directly -- all 4 were already
    present (黑龙江's natively via celma's own crawl, 内蒙古's via an earlier
    Wind-reconcile patch), just unmatched by the "债"-only normalization.
    Before this fix the SSE gap count over-reported by up to 4 whenever
    these two provinces had recent listings -- **always check a
    "missing" bond against state_results.csv by eye for 内蒙古/黑龙江 before
    trusting this function's output for them**, or better, keep this alias
    map current if a third such province is ever found."""
    start = _months_back_start(months_back)
    end = datetime.date.today().isoformat()
    notices = sse_listing.fetch_recent_listing_notices(start, end)

    _PROVINCE_ROOT_ALIASES = {"黑龙江": "龙江", "内蒙古": "内蒙"}

    def _normalize(s: str) -> str:
        s = s.replace("债", "")
        for full, short in _PROVINCE_ROOT_ALIASES.items():
            s = s.replace(full, short)
        return s

    known_shortnames = set(state_df["bond_short_name"].dropna())
    known_normalized = {_normalize(s) for s in known_shortnames}

    def _is_known(abbr: str) -> bool:
        return abbr in known_shortnames or _normalize(abbr) in known_normalized

    missing = [
        n for n in notices
        if n.get("securityAbbr") and not _is_known(n["securityAbbr"])
    ]
    by_province: dict[str, dict] = {}
    for n in notices:
        prov = n.get("province") or "未知"
        by_province.setdefault(prov, {"checked": 0, "missing": 0})
        by_province[prov]["checked"] += 1
        if n.get("securityAbbr") and not _is_known(n["securityAbbr"]):
            by_province[prov]["missing"] += 1
    return {"checked": len(notices), "missing_from_celma": missing, "missing_by_province": by_province}


def run_validation(months_back: int = 2, state_results_path=None) -> dict:
    """Runs both cross-checks (SSE existence check, provincial full
    diff_against_state per province) against the CURRENT on-disk
    state_results.csv. Returns a dict with everything a report generator or
    apply_corrections() needs -- does not mutate state_results.csv itself."""
    state_results_path = state_results_path or config.STATE_RESULTS_CSV
    state_df = pd.read_csv(state_results_path)

    sse_result = check_sse_coverage(state_df, months_back=months_back)

    provincial_rows_by_province = provincial_crawl.crawl_all(months_back=months_back)
    provincial_diffs = {}
    for province, rows in provincial_rows_by_province.items():
        provincial_diffs[province] = provincial_verify.diff_against_state(rows, state_df)

    return {
        "run_at": datetime.datetime.now().isoformat(timespec="seconds"),
        "months_back": months_back,
        "state_row_count": len(state_df),
        "sse": sse_result,
        "provincial": provincial_diffs,
    }


def _unique_candidate_mask(state_df: pd.DataFrame, row: dict) -> pd.DataFrame:
    if row.get("bond_short_name"):
        hit = state_df[
            (state_df["province"] == row["province"])
            & (state_df["bond_short_name"] == row["bond_short_name"])
        ]
        if len(hit) == 1:
            return hit
    candidates = state_df[
        (state_df["province"] == row["province"])
        & (state_df["issue_date"] == row["issue_date"])
        & (state_df["term"] == row["term"])
    ]
    return candidates if len(candidates) == 1 else candidates.iloc[0:0]


def apply_corrections(validation: dict, state_results_path=None, dry_run: bool = False) -> dict:
    """Patches state_results.csv's total_amount_yi/coupon_rate_pct where a
    provincial source disagrees with celma on a row it matched UNAMBIGUOUSLY
    (see module docstring for the exact conditions -- deliberately more
    conservative than wind_reconcile.py's auto-apply). Never adds rows.
    Idempotent: a second run against an already-corrected CSV should apply
    zero further changes for the same underlying data."""
    state_results_path = state_results_path or config.STATE_RESULTS_CSV
    state_df = pd.read_csv(state_results_path)

    fixed_amount = fixed_rate = 0
    applied = []
    for province, diff in validation["provincial"].items():
        for entry in diff["amount_mismatch"] + diff["rate_mismatch"]:
            row = entry["row"]
            candidates = _unique_candidate_mask(state_df, row)
            if candidates.empty:
                continue  # ambiguous even at apply-time (state_df may have shifted) -- skip, don't guess
            i = candidates.index[0]
            note = []
            if row.get("total_amount_yi") is not None:
                old = state_df.at[i, "total_amount_yi"]
                if abs(old - row["total_amount_yi"]) > AMOUNT_TOLERANCE_YI:
                    state_df.at[i, "total_amount_yi"] = row["total_amount_yi"]
                    note.append(f"total_amount_yi由{old}订正为{row['total_amount_yi']}(省级财政厅公告核对)")
                    fixed_amount += 1
            if row.get("coupon_rate_pct") is not None:
                old = state_df.at[i, "coupon_rate_pct"]
                if pd.isna(old) or abs(old - row["coupon_rate_pct"]) > RATE_TOLERANCE_PCT:
                    state_df.at[i, "coupon_rate_pct"] = row["coupon_rate_pct"]
                    note.append(f"coupon_rate_pct由{old}订正为{row['coupon_rate_pct']}(省级财政厅公告核对)")
                    fixed_rate += 1
            if note:
                existing_warn = state_df.at[i, "warnings"]
                new_warn = "; ".join(note)
                state_df.at[i, "warnings"] = f"{existing_warn}; {new_warn}" if isinstance(existing_warn, str) and existing_warn else new_warn
                applied.append({"province": province, "bond_short_name": state_df.at[i, "bond_short_name"], "notes": note})

    if not dry_run and (fixed_amount or fixed_rate):
        state_df = state_df.sort_values("pub_date").reset_index(drop=True)
        state_df.to_csv(state_results_path, index=False)

    return {"amount_fixed": fixed_amount, "rate_fixed": fixed_rate, "applied": applied}


# ---------------------------------------------------------------------------
# 校验报表 (Excel validation report)
# ---------------------------------------------------------------------------

_DIFF_CATEGORY_LABELS = {
    "matched": "一致确认", "amount_mismatch": "发行规模不一致",
    "rate_mismatch": "票面利率不一致", "not_found_in_state": "celma未收录",
    "low_confidence": "低置信度(疑似OCR串块,需人工核对)",
}


def build_validation_report_xlsx(validation: dict, corrections: dict | None = None, out_path=None):
    """Writes the 三方交叉校验报表 (celma vs SSE vs 省级财政厅公告) to
    output/provincial_validation_report.xlsx. Every row a human would need
    to manually check is listed with its full detail (province, source
    announcement URL, both sides' figures) -- summary counts alone aren't
    enough to act on."""
    out_path = out_path or (config.OUTPUT_DIR / "provincial_validation_report.xlsx")

    overview_rows = [
        {"项目": "运行时间", "值": validation["run_at"]},
        {"项目": "回溯月数", "值": validation["months_back"]},
        {"项目": "state_results.csv 总行数", "值": validation["state_row_count"]},
        {"项目": "SSE上市公告核对数", "值": validation["sse"]["checked"]},
        {"项目": "SSE已确认但celma未收录", "值": len(validation["sse"]["missing_from_celma"])},
    ]
    for prov, diff in validation["provincial"].items():
        for cat, label in _DIFF_CATEGORY_LABELS.items():
            n = len(diff[cat])
            if n:
                overview_rows.append({"项目": f"{prov} - {label}", "值": n})
    if corrections:
        overview_rows.append({"项目": "本次已自动修正-发行规模", "值": corrections["amount_fixed"]})
        overview_rows.append({"项目": "本次已自动修正-票面利率", "值": corrections["rate_fixed"]})
    overview_df = pd.DataFrame(overview_rows)

    sse_by_province_rows = [
        {"省份": prov, "SSE核对数": d["checked"], "celma未收录数": d["missing"],
         "未收录占比": f"{100*d['missing']/d['checked']:.0f}%" if d["checked"] else "0%"}
        for prov, d in sorted(validation["sse"]["missing_by_province"].items(),
                               key=lambda kv: -kv[1]["missing"])
    ]
    sse_province_df = pd.DataFrame(sse_by_province_rows)
    sse_missing_df = pd.DataFrame(validation["sse"]["missing_from_celma"])

    detail_rows = []
    for prov, diff in validation["provincial"].items():
        for cat, label in _DIFF_CATEGORY_LABELS.items():
            for entry in diff[cat]:
                row = entry["row"]
                srow = entry.get("state_row", {})
                detail_rows.append({
                    "省份": prov, "核对结论": label,
                    "省级公告债券名称": row.get("bond_name"),
                    "省级公告发行日期": row.get("issue_date"),
                    "省级公告期限": row.get("term"),
                    "省级公告发行规模(亿)": row.get("total_amount_yi"),
                    "省级公告票面利率(%)": row.get("coupon_rate_pct"),
                    "celma债券简称": srow.get("bond_short_name"),
                    "celma发行规模(亿)": srow.get("total_amount_yi"),
                    "celma票面利率(%)": srow.get("coupon_rate_pct"),
                    "原因/备注": row.get("warnings") or entry.get("reason") or "",
                    "公告来源URL": row.get("source_url"),
                })
    detail_df = pd.DataFrame(detail_rows)

    applied_df = pd.DataFrame(corrections["applied"]) if corrections and corrections.get("applied") else pd.DataFrame()

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        overview_df.to_excel(writer, sheet_name="总览", index=False)
        sse_province_df.to_excel(writer, sheet_name="SSE缺口(按省份)", index=False)
        if len(sse_missing_df):
            sse_missing_df.to_excel(writer, sheet_name="SSE缺口明细", index=False)
        if len(detail_df):
            detail_df.to_excel(writer, sheet_name="省级财政厅核对明细", index=False)
        if len(applied_df):
            applied_df.to_excel(writer, sheet_name="已自动修正", index=False)

    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
        for i, col in enumerate(ws.iter_cols(1, ws.max_column), start=1):
            from openpyxl.utils import get_column_letter
            header_len = len(str(col[0].value or ""))
            ws.column_dimensions[get_column_letter(i)].width = max(14, min(header_len + 4, 45))
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# chinabond.com.cn gap-fill check (on-demand, NOT part of run_validation())
# ---------------------------------------------------------------------------
#
# Deliberately NOT wired into run_validation()/main.py's default path:
# src/chinabond_crawl.py's OCR pass is expensive (confirmed 2026-08-12: 10
# provinces took well over 2 minutes and hit a transient proxy failure once)
# and, per that module's docstring, a raw not_found count from it isn't
# trustworthy on its own -- every "not found" row where the parser also
# failed to extract issue_date/term needs a human glance at the PDF, not an
# automatic verdict. So this is an on-demand tool for the specific moment
# you already know (from check_sse_coverage()) which bonds are missing and
# want to check whether chinabond.com.cn happens to have the confirming
# document -- not a step that should silently run on every pipeline build.

def check_chinabond_for_sse_gaps(sse_missing: list[dict], page_size: int = 10) -> list[dict]:
    """For each SSE-flagged missing bond (as returned by
    check_sse_coverage()['missing_from_celma']), fetch that province's
    recent chinabond.com.cn 发行结果 items and report whether any of them
    plausibly cover it (by date + a loose title match on the bond's
    tranche description). Returns a list of {"sse_bond", "chinabond_match"}
    -- `chinabond_match` is None if nothing plausible was found in the
    fetched page. This does NOT parse/confirm exact amount/rate; that
    still requires reading the actual PDF (see chinabond_crawl.py's module
    docstring for why automated extraction isn't reliable enough to trust
    blindly here)."""
    from . import chinabond_crawl as cb

    by_province: dict[str, list[dict]] = {}
    results = []
    for bond in sse_missing:
        prov = bond.get("province")
        if not prov:
            results.append({"sse_bond": bond, "chinabond_match": None})
            continue
        if prov not in by_province:
            try:
                by_province[prov] = cb.fetch_channel_page(cb.CHANNEL_RESULTS, issuer=prov, page_size=page_size)
            except Exception:
                by_province[prov] = []
        candidates = [
            item for item in by_province[prov]
            if item.get("createTime") and bond.get("sseDate")
            and abs((pd.Timestamp(item["createTime"]) - pd.Timestamp(bond["sseDate"])).days) <= 3
        ]
        results.append({
            "sse_bond": bond,
            "chinabond_match": candidates[0] if candidates else None,
        })
    return results
